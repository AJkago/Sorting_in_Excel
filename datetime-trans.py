import openpyxl
import datetime

file_name = r'C:\Users\chika\花_180401_210331.xlsx'

wb = openpyxl.load_workbook(file_name)
sheet0 = wb['Sheet0']
sheet1 = wb.copy_worksheet(sheet0)
sheet1.title = 'Sheet1'

lastrow = sheet0.max_row
print(lastrow)

i = 2
for row in sheet0.iter_rows():
    for cell in row:
        if cell.col_idx == 1:
            for i in range(2, lastrow+1):
                betauchi = sheet0.cell(i,1).value
                if betauchi is None:
                    break
                outlist = []
                outlist = list(betauchi)
                print(outlist)
                seireki8 = []
                j =  0
                for j in range(0,len(outlist)):
                    if 0<=j<4:
                        seireki8.append(outlist[j])
                    elif j==8:
                        if outlist[j] == '日':
                            seireki8.append('0')
                            seireki8.append(outlist[j-3])
                            seireki8.append('0')
                            seireki8.append(outlist[j-1])
                            break
                        elif outlist[j-2] == '月':
                            seireki8.append('0')
                            seireki8.append(outlist[j-3])
                            seireki8.append(outlist[j-1])
                            seireki8.append(outlist[j])
                            break
                    elif j==9:
                        if outlist[j] == '日':
                            seireki8.append(outlist[j-4])
                            seireki8.append(outlist[j-3])
                            seireki8.append('0')
                            seireki8.append(outlist[j-1])
                        else:
                            seireki8.append(outlist[j-4])
                            seireki8.append(outlist[j-3])
                            seireki8.append(outlist[j-1])
                            seireki8.append(outlist[j])
                    j = j + 1

                seireki8_str = ''.join(seireki8)
                print(type(seireki8_str))
                shukkabi = datetime.datetime.strptime(seireki8_str,'%Y%m%d')
                sheet1.cell(row=i,column=1,value = shukkabi)
                i = i + 1
        break
    break
#保存
wb.save(file_name)
