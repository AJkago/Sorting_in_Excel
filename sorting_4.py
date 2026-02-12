import openpyxl

file_name = r'C:\Users\chika\2021切花.xlsx'

wb = openpyxl.load_workbook(file_name)
ws3 = wb['Sheet3']
ws4 = wb.create_sheet(title='Sheet4',index=4)

lastrow = ws3.max_row
lastcolumn = ws3.max_column

#配列宣言
datecolumns = []

i = 1
num = 2
for row in ws3.iter_rows(max_col=lastcolumn):
    for cell in row:
        if cell.row == 1:
            if i == 1:
                k = 2
                for k in range(2,lastrow+1):
                    filled_cell = ws3.cell(row = k, column = 1).value
                    print(filled_cell)
                    if filled_cell is not None:
                        ws4.cell(row = k, column = 1).value = filled_cell
                        k = k + 1
                i += 1
            elif i == 2:
                datecolumns.append(ws3.cell(row=1,column=i).value)
                k = 2
                for k in range(2,lastrow+1):
                    filled_cell = ws3.cell(row = k, column = 2).value
                    print(filled_cell)
                    if filled_cell is not None:
                        ws4.cell(row = k, column = 2).value = filled_cell
                        k = k + 1
                i += 1
            else:
                for list in datecolumns:
                    j = 0
                    if list == ws3.cell(row=1, column=i).value:
                        j = j + 1
                        break
                if j == 0:
                    if ws3.cell(row=1, column=i).value is not None:
                        datecolumns.append(ws3.cell(row=1, column=i).value)
                        num = num + 1
                k = 2
                new_cell = 0
                for k in range(2,lastrow+1):
                    filled_cell = ws3.cell(row = k, column = i).value
                    if filled_cell is None:
                        filled_cell = 0
                    filled_cell_int = int(filled_cell)

                    new_cell = ws4.cell(row = k, column = num).value
                    if new_cell is None:
                        new_cell = 0
                    new_cell_int = int(new_cell)
                    ws4.cell(row = k, column = num).value = new_cell_int + filled_cell_int
                    k = k + 1
                i += 1
    break
#重複しないリストをws3へ書き込み
print(len(datecolumns))
for i in range(0,len(datecolumns)):
    ws4.cell(1,i+2,value = datecolumns[i])

print(lastcolumn)

#保存
wb.save(file_name)
