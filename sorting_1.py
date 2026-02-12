import openpyxl

file_name = r'C:\Users\chika\2021切花.xlsx'

wb = openpyxl.load_workbook(file_name)
sheet1 = wb['Sheet1']
sheet2 = wb.create_sheet(title='Sheet2',index=2)

rw = sheet1.max_row

for i in range(1,rw):
    c_a_1 = sheet1.cell(row=i+1,column=1)
    r_2_2 = sheet2.cell(row=1,column=i+1)
    r_2_2.value = c_a_1.value

    c_b_1 = sheet1.cell(row=i+1,column=2)
    c_b_2 = sheet2.cell(row=i+1,column=1)
    c_b_2.value = c_b_1.value

    c_c_1 = sheet1.cell(row=i+1,column=3)
    r_c_x = sheet2.cell(row=i+1,column=i+1)
    r_c_x.value   = c_c_1.value

wb.save(file_name)

print(sheet2)
