import openpyxl

file_name = r'C:\Users\chika\2021切花.xlsx'

wb = openpyxl.load_workbook(file_name)
ws2 = wb['Sheet2']
ws3 = wb.create_sheet(title='Sheet3',index=3)

lastcolumn = ws2.max_column

#配列宣言
Gender = []

#重複しないリストを抽出
#セルをループ
i = 1
num = 1
for row in ws2.iter_rows():
    j = 0
    for cell in row:
        if cell.col_idx == 1:
            if i == 1:
            #A1は強制的に配列へ格納
                Gender.append(cell.value)
            else:
                #リストをループ
                for list in Gender:
                    #セル値がすでにリストに含まれていたら何もしない
                    if list == cell.value:
                        j = j + 1
                        break
                if j == 0:
                    #セル値がリストに含まれていなければ配列へ追加
                    if cell.value is not None:
                        Gender.append(cell.value)
                        num = num + 1
            k = 1
            for k in range(1,lastcolumn+1):
                filled_cell = ws2.cell(row = i, column = k).value
                if filled_cell is not None:
                    ws3.cell(row = num, column = k).value = filled_cell
                    k = k + 1
            i = i + 1

#重複しないリストをws3へ書き込み
for i in range(0,len(Gender)):
    ws3.cell(i+1,1,value = Gender[i])

#保存
wb.save(file_name)
